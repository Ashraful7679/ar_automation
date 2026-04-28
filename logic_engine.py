import sqlite3
import csv
import openpyxl
import os
import xlrd
from pypdf import PdfReader
import re
import pandas as pd
from parsers.arabian_shield import ArabianShieldParser
from parsers.axa_ppp import AxaPppParser
from parsers.msh import MshParser
from parsers.healix import HealixParser
from parsers.sos import SosParser
from parsers.gems import GemsParser
from parsers.payadvice import PayAdviceParser
from parsers.worldwide import WorldwideParser
from parsers.nextcare import NextcareParser
from parsers.health360 import Health360Parser
from fpdf import FPDF
from datetime import datetime
import sys

class LogicEngine:
    def __init__(self, db_path=":memory:"):
        self.db_path = db_path
        self.conn = sqlite3.connect(db_path, timeout=30)
        self.conn.row_factory = sqlite3.Row
        self.cursor = self.conn.cursor()
        self.table_name = "raw_data"
        self.formatted_table_name = "formatted_data"
        self.headers = []
        self.formatted_headers = []
        self.current_parser = None

        self._recover_raw_headers()
        self._recover_formatted_headers()

    def _get_base_path(self):
        if getattr(sys, 'frozen', False):
            return sys._MEIPASS
        return os.path.dirname(os.path.abspath(__file__))

    def _recover_formatted_headers(self):
        try:
            self.cursor.execute(f"PRAGMA table_info({self.formatted_table_name})")
            cols = self.cursor.fetchall()
            if cols:
                self.formatted_headers = [c[1] for c in cols]
        except Exception:
            pass

    def _recover_raw_headers(self):
        try:
            self.cursor.execute(f"PRAGMA table_info({self.table_name})")
            cols = self.cursor.fetchall()
            if cols:
                self.headers = [c[1] for c in cols]
        except Exception:
            pass

    def _get_parser(self, profile_name):
        if profile_name == 'ARABIAN_SHIELD':
            return ArabianShieldParser()
        elif profile_name == 'AXA_PPP':
            return AxaPppParser()
        elif profile_name == 'MSH':
            return MshParser()
        elif profile_name == 'HEALIX':
            return HealixParser()
        elif profile_name == 'SOS':
            return SosParser()
        elif profile_name == 'GEMS':
            return GemsParser()
        elif profile_name == 'PAYADVICE':
            return PayAdviceParser()
        elif profile_name == 'TAWUNIYA':
            return WorldwideParser() # Placeholder
        elif profile_name == 'BUPA':
            return WorldwideParser() # Placeholder
        elif profile_name == 'WORLDWIDE':
            return WorldwideParser()
        elif profile_name == 'NEXT_CARE':
            return NextcareParser()
        elif profile_name == 'HEALTH_360':
            return Health360Parser()
        return None

    def _run_parser(self, parser_instance, file_path):
        self.current_parser = parser_instance
        try:
            rows = self.current_parser.parse(file_path)
        except Exception as e:
            print(f"Parser Error: {e}")
            raise e
        
        self.headers = self.current_parser.raw_headers
        self._create_table(self.headers)
        if rows:
            # Pad rows to match the new headers (which include the extra columns)
            num_cols = len(self.headers)
            padded_rows = [list(r) + [""] * (num_cols - len(r)) for r in rows]
            
            placeholder = ','.join(['?']*num_cols)
            cols = ', '.join([f'"{h}"' for h in self.headers])
            self.cursor.executemany(f"INSERT INTO {self.table_name} ({cols}) VALUES ({placeholder})", padded_rows)
        self.conn.commit()

    def _create_formatted_table(self, headers):
        cols = [f'"{h}" TEXT' for h in headers]
        self.cursor.execute(f"DROP TABLE IF EXISTS {self.formatted_table_name}")
        self.cursor.execute(f"CREATE TABLE {self.formatted_table_name} ({', '.join(cols)})")

    def _persist_formatted_data(self):
        if not self.current_parser:
             print("DEBUG: _persist_formatted_data: No current_parser set!")
             return
        print(f"DEBUG: _persist_formatted_data: Using parser {type(self.current_parser).__name__}")

        self.cursor.execute(f'SELECT * FROM {self.table_name}')
        col_names = [desc[0] for desc in self.cursor.description]
        if 'id' in col_names:
            col_names.remove('id')

        self.cursor.execute(f"SELECT {', '.join([f'\"{c}\"' for c in col_names])} FROM {self.table_name}")
        raw_rows = self.cursor.fetchall()

        self.formatted_headers, formatted_rows = self.current_parser.transform(raw_rows)

        self._create_formatted_table(self.formatted_headers)

        if formatted_rows:
            placeholder = ','.join(['?']*len(self.formatted_headers))
            cols = ', '.join([f'"{h}"' for h in self.formatted_headers])
            self.cursor.executemany(f"INSERT INTO {self.formatted_table_name} ({cols}) VALUES ({placeholder})", formatted_rows)
            print(f"DEBUG: _persist_formatted_data: Inserted {len(formatted_rows)} rows into {self.formatted_table_name}")
        self.conn.commit()

    def load_file(self, file_path, profile_name='default', sheet_name=None):
        ext = os.path.splitext(file_path)[1].lower()

        self.cursor.execute(f"DROP TABLE IF EXISTS {self.table_name}")

        try:
            if ext == '.pdf':
                self.current_parser = self._get_parser(profile_name)
                if self.current_parser:
                    self._run_parser(self.current_parser, file_path)
                    self._persist_formatted_data()
                else:
                    self._load_pdf(file_path)
            elif ext == '.csv':
                self._load_csv(file_path)
            elif ext in ['.xlsx', '.xlsm']:
                self._load_excel(file_path, sheet_name=sheet_name)
            elif ext == '.xls':
                self._load_xls(file_path, sheet_name=sheet_name)
            else:
                raise ValueError(f"Unsupported file type: {ext}")
        finally:
            if os.path.exists(file_path) and "uploads" in file_path:
                try:
                    os.remove(file_path)
                    print(f"DEBUG: Deleted uploaded file {file_path}")
                except Exception as e:
                    print(f"WARNING: Could not delete uploaded file {file_path}: {e}")

    def get_formatted_preview(self):
        # Always try to fetch from formatted_table first
        try:
            placeholders = ', '.join([f'"{h}"' for h in self.formatted_headers])
            self.cursor.execute(f"SELECT rowid, {placeholders} FROM {self.formatted_table_name}")
            rows = self.cursor.fetchall()
            
            if rows:
                return {
                    "headers": ["_id"] + self.formatted_headers,
                    "rows": [list(r) for r in rows]
                }
        except Exception as e:
            print(f"DEBUG: get_formatted_preview error: {e}")
        
        # Fallback to parser/raw only if formatted table is EMPTY
        return self.get_preview()

    FIXED_HEADERS = ["Sl. No", "Inv No", "Date", "Patient ID", "Patient Name", "Invoice Balance", "Amt To Adjust", "CustomerCode", "Remark"]

    def generate_outputs(self, profile_name='default'):
        """
        Main entry point for "Run Process".
        Splits data by CustomerCode and generates separate files.
        """
        print(f"DEBUG: generate_outputs called for profile: {profile_name}")
        
        # Ensure we have data
        self.cursor.execute(f"SELECT COUNT(*) FROM {self.formatted_table_name}")
        count = self.cursor.fetchone()[0]
        if count == 0:
            print("DEBUG: generate_outputs: Table is empty, nothing to process.")
            return []

        if not self.formatted_headers:
            self._recover_formatted_headers()

        # Find CustomerCode index
        cc_idx = -1
        try:
            cc_idx = self.formatted_headers.index("CustomerCode")
        except ValueError:
            print("WARNING: CustomerCode column not found in headers")
            filename = self.generate_custom_output(profile_name=profile_name, file_format='xlsx')
            return [filename]

        # Get unique CustomerCodes
        self.cursor.execute(f"SELECT DISTINCT \"CustomerCode\" FROM {self.formatted_table_name}")
        codes = [r[0] for r in self.cursor.fetchall() if r[0]]
        
        if not codes:
            filename = self.generate_custom_output(profile_name=profile_name, file_format='xlsx')
            return [filename]

        generated_files = []
        for code in codes:
            filename = self.generate_custom_output(
                profile_name=profile_name, 
                custom_filename=f"MathingOfARReceipts_{code}",
                file_format='xls',
                customer_code=code
            )
            generated_files.append(filename)

        return generated_files

    def generate_custom_output(self, profile_name=None, custom_filename=None, file_format='xlsx', customer_code=None):
        print(f"DEBUG: generate_custom_output called with profile={profile_name}, filename={custom_filename}, format={file_format}, customer_code={customer_code}")
        
        if getattr(sys, 'frozen', False):
             base_dir = os.path.dirname(sys.executable)
             output_dir = os.path.join(base_dir, 'output')
        else:
             output_dir = os.path.join(self._get_base_path(), 'output')
        
        if not os.path.exists(output_dir):
             os.makedirs(output_dir)

        ext = file_format.lower()
        if ext not in ['xlsx', 'csv', 'pdf', 'xls', 'xlsm']:
            ext = 'xls'
        
        if not custom_filename:
            base_name = f"Formatted_Output_{profile_name}" if profile_name else "Processed_Output"
            filename = f"{base_name}.{ext}"
        else:
            # Remove any existing extension and re-add correct one
            base_name = custom_filename.rsplit('.', 1)[0]
            filename = f"{base_name}.{ext}"
            
        output_path = os.path.join(output_dir, filename)

        # Overwrite if exists (no unique names requirement)
        if os.path.exists(output_path):
             try:
                 os.remove(output_path)
             except Exception:
                 pass

        if not self.formatted_headers:
            self._recover_formatted_headers()

        if not self.formatted_headers:
            self.formatted_headers = self.FIXED_HEADERS

        query_cols = ', '.join([f'"{h}"' for h in self.formatted_headers])

        try:
            if customer_code:
                print(f"DEBUG: Filtering data for CustomerCode: {customer_code}")
                self.cursor.execute(f"SELECT {query_cols} FROM {self.formatted_table_name} WHERE \"CustomerCode\" = ?", (customer_code,))
            else:
                self.cursor.execute(f"SELECT {query_cols} FROM {self.formatted_table_name}")
            
            raw_rows = self.cursor.fetchall()
            # Convert sqlite3.Row objects to regular lists
            formatted_rows = [list(row) for row in raw_rows]
            formatted_headers = list(self.formatted_headers)
            print(f"DEBUG: generate_custom_output: fetched {len(formatted_rows)} rows from DB")

            # Apply ERP compatibility cleaning
            cleaned_rows = []
            for row in formatted_rows:
                new_row = list(row)
                # Force Sl. No to float if numeric
                try:
                    if new_row[0]: new_row[0] = float(str(new_row[0]).strip())
                except: pass

                # Clean strings: remove Hex artifacts and normalize line breaks
                for i in range(len(new_row)):
                    if isinstance(new_row[i], str):
                        # Strip _x005F_x000D_ and internal newlines
                        new_row[i] = new_row[i].replace('_x005F_x000D_', ' ')
                        new_row[i] = new_row[i].replace('\r', ' ').replace('\n', ' ')
                        new_row[i] = ' '.join(new_row[i].split()) # Remove double spaces
                cleaned_rows.append(new_row)
            formatted_rows = cleaned_rows

            # Merge Remark 1, 2, 3 into a single Remark column (11 cols -> 9 cols)
            if len(formatted_headers) == 11 and formatted_headers[8] == "Remark 1":
                new_headers = formatted_headers[:8] + ["Remark"]
                new_rows = []
                for r in formatted_rows:
                    new_r = r[:8]
                    # Combine columns 8, 9, 10, avoiding 'None' strings and empty values
                    remarks = []
                    for x in r[8:]:
                        val = str(x).strip() if x is not None else ""
                        if val and val.lower() != "none":
                            remarks.append(val)
                    new_r.append(", ".join(remarks))
                    new_rows.append(new_r)
                formatted_headers = new_headers
                formatted_rows = new_rows
        except Exception as e:
            print(f"DEBUG: Exception fetching formatted data: {e}")
            formatted_headers = list(self.formatted_headers)
            formatted_rows = []

        if ext == 'csv':
            with open(output_path, 'w', newline='', encoding='utf-8-sig') as f:
                # Force CRLF and Quoting for ERP compatibility
                writer = csv.writer(f, lineterminator='\r\n', quoting=csv.QUOTE_ALL)
                for _ in range(4):
                    writer.writerow([])
                writer.writerow(formatted_headers)
                # Format numeric columns to 3 decimals for CSV
                rows_to_csv = []
                for row in formatted_rows:
                    new_row = list(row)
                    for i in [5, 6]:
                        if i < len(new_row) and new_row[i] is not None:
                            try:
                                new_row[i] = f"{float(new_row[i]):.3f}"
                            except:
                                pass
                    rows_to_csv.append(new_row)
                writer.writerows(rows_to_csv)
        elif ext == 'pdf':
            self._generate_pdf(output_path, formatted_headers, formatted_rows, profile_name)
        elif ext == 'xls':
            try:
                # Prepare data with pandas first for numeric cleaning
                df = pd.DataFrame(formatted_rows, columns=formatted_headers)
                for i in [5, 6]:
                    if i < len(formatted_headers):
                        col_name = formatted_headers[i]
                        df[col_name] = pd.to_numeric(df[col_name].astype(str).str.replace(',', '').str.strip(), errors='coerce')
                
                rows_to_write = df.values.tolist()
                
                template_path = os.path.join(self._get_base_path(), 'templates', 'output.xls')
                print(f"DEBUG: template_path={template_path}, exists={os.path.exists(template_path)}")
                if os.path.exists(template_path):
                    try:
                        import xlrd
                        from xlutils.copy import copy
                        # formatting_info=True allows preserving styles during copy
                        rb = xlrd.open_workbook(template_path, formatting_info=True)
                        wb = copy(rb)
                        ws = wb.get_sheet(0)
                        
                        # Write data starting from A6 (Row Index 5)
                        import xlwt
                        num_style = xlwt.easyxf(num_format_str='0.000')
                        for r_idx, row_data in enumerate(rows_to_write):
                            for c_idx, val in enumerate(row_data):
                                if c_idx < 9: # Only A-I (9 columns)
                                    if c_idx in [5, 6]:
                                        ws.write(r_idx + 5, c_idx, val, num_style)
                                    else:
                                        ws.write(r_idx + 5, c_idx, val)
                        wb.save(output_path)
                    except Exception as e:
                        print(f"DEBUG: xlutils copy failed: {e}")
                        # Fallback
                        with pd.ExcelWriter(output_path, engine='xlwt') as writer:
                            writer.book.owner = "dell"
                            df.to_excel(writer, index=False, startrow=4, sheet_name='Sheet1')
                else:
                    # Fallback to standard creation if template missing
                    print(f"DEBUG: Template not found at {template_path}, using standard creation")
                    with pd.ExcelWriter(output_path, engine='xlwt') as writer:
                        writer.book.owner = "dell"
                        df.to_excel(writer, index=False, startrow=4, sheet_name='Sheet1')
            except Exception as e:
                print(f"Failed to export XLS: {e}")

        elif ext in ['xlsx', 'xlsm']:
            from openpyxl.styles import numbers

            # Use pandas for faster data processing if possible
            try:
                print(f"DEBUG: generate_custom_output: Creating DataFrame with {len(formatted_rows)} rows")
                df = pd.DataFrame(formatted_rows, columns=formatted_headers)
                # Convert Invoice Balance and Amt To Adjust to numbers (indices 5 and 6)
                for i in [5, 6]:
                    col_name = formatted_headers[i]
                    df[col_name] = pd.to_numeric(df[col_name].astype(str).str.replace(',', '').str.strip(), errors='coerce')

                # Write to Excel
                with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                    df.to_excel(writer, index=False, startrow=4, sheet_name='Data')
                    
                    # Apply numeric formatting to columns F and G
                    ws = writer.sheets['Data']
                    for row in ws.iter_rows(min_row=6, max_row=ws.max_row, min_col=6, max_col=7):
                        for cell in row:
                            if cell.value is not None:
                                cell.number_format = '0.000'
            except Exception as e:
                # Fallback to manual openpyxl if pandas fails
                print(f"Pandas export failed, using fallback: {e}")
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Data"
                for _ in range(4): ws.append([])
                ws.append(formatted_headers)
                for r in formatted_rows:
                    new_row = list(r)
                    for i in [5, 6]:
                        try:
                            val = str(new_row[i]).replace(',', '').strip()
                            if val: new_row[i] = float(val)
                        except: pass
                    ws.append(new_row)
                
                for row in ws.iter_rows(min_row=6, max_row=ws.max_row, min_col=6, max_col=7):
                    for cell in row:
                        if isinstance(cell.value, (int, float)):
                            cell.number_format = numbers.FORMAT_NUMBER_00
                wb.save(output_path)
                print(f"DEBUG: XLS file saved using template at {template_path}")

        print(f"DEBUG: returning filename={filename}")
        return filename

    def _generate_pdf(self, output_path, headers, rows, profile_name):
        pdf = FPDF(orientation='L', unit='mm', format='A4')
        pdf.add_page()

        logo_path = os.path.join(self._get_base_path(), 'static', 'img', 'logo.png')
        if os.path.exists(logo_path):
            pdf.image(logo_path, 10, 8, 33)

        pdf.set_font('helvetica', 'B', 20)
        pdf.set_text_color(31, 41, 55)
        pdf.cell(80)
        pdf.cell(100, 10, 'AR Report', 0, 1, 'L')

        pdf.set_font('helvetica', '', 10)
        pdf.set_text_color(107, 114, 128)
        pdf.cell(80)
        pdf.cell(100, 5, f'Generated on: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}', 0, 1, 'L')
        pdf.cell(80)
        pdf.cell(100, 5, f'Payer Profile: {profile_name or "General"}', 0, 1, 'L')
        pdf.ln(20)

        pdf.set_font('helvetica', 'B', 9)
        pdf.set_fill_color(243, 244, 246)
        pdf.set_text_color(0, 0, 0)

        widths = [12, 35, 25, 30, 60, 25, 25, 35, 30]

        for i, h in enumerate(headers):
            pdf.cell(widths[i], 10, str(h), 1, 0, 'C', 1)
        pdf.ln()

        pdf.set_font('helvetica', '', 8)
        total_bal = 0
        total_adj = 0

        for r in rows:
            if pdf.get_y() > 175:
                pdf.add_page()
                pdf.set_font('helvetica', 'B', 9)
                pdf.set_fill_color(243, 244, 246)
                for i, h in enumerate(headers):
                    pdf.cell(widths[i], 10, str(h), 1, 0, 'C', 1)
                pdf.ln()
                pdf.set_font('helvetica', '', 8)

            for i, cell in enumerate(r):
                val = str(cell) if cell is not None else ""
                align = 'L'
                if i in [0, 5, 6]: align = 'R'

                max_w = widths[i] - 2
                if pdf.get_string_width(val) > max_w:
                    while pdf.get_string_width(val + "...") > max_w and len(val) > 0:
                        val = val[:-1]
                    val = val + "..."

                pdf.cell(widths[i], 8, val, 1, 0, align)

            try:
                bal_str = str(r[5]).replace(',', '') if r[5] else '0'
                adj_str = str(r[6]).replace(',', '') if r[6] else '0'
                total_bal += float(bal_str) if bal_str else 0
                total_adj += float(adj_str) if adj_str else 0
            except: pass

            pdf.ln()

        pdf.set_font('helvetica', 'B', 9)
        pdf.set_fill_color(249, 250, 251)
        pdf.cell(widths[0], 10, "", 1, 0, 'C', 1)
        pdf.cell(widths[1] + widths[2] + widths[3] + widths[4], 10, "GRAND TOTALS", 1, 0, 'R', 1)
        pdf.cell(widths[5], 10, f"{total_bal:,.2f}", 1, 0, 'R', 1)
        pdf.cell(widths[6], 10, f"{total_adj:,.2f}", 1, 0, 'R', 1)
        pdf.cell(widths[7] + widths[8], 10, "", 1, 0, 'C', 1)
        pdf.ln(15)

        pdf.set_font('helvetica', 'I', 8)
        pdf.set_text_color(156, 163, 175)
        pdf.cell(0, 10, "Note: This is a system-generated report. Confidential.", 0, 1, 'C')
        pdf.cell(0, 5, f"Verified by: ____________________   Date: {datetime.now().strftime('%Y-%m-%d')}", 0, 1, 'R')

        pdf.output(output_path)

    def update_formatted_cell(self, row_id, col_index, new_value):
        if not self.formatted_headers:
             self._recover_formatted_headers()

        if not self.formatted_headers:
             return False

        if col_index < 0 or col_index >= len(self.formatted_headers):
            raise ValueError("Invalid column index")

        col_name = self.formatted_headers[col_index]
        self.cursor.execute(f'UPDATE {self.formatted_table_name} SET "{col_name}" = ? WHERE rowid = ?', (new_value, row_id))
        self.conn.commit()
        return True

    def update_formatted_cells_batch(self, updates):
        if not self.formatted_headers:
             self._recover_formatted_headers()

        if not self.formatted_headers:
             return False

        updates_by_col = {}
        for u in updates:
            idx = int(u['col_index'])
            if idx < 0 or idx >= len(self.formatted_headers): continue
            col_name = self.formatted_headers[idx]

            if col_name not in updates_by_col:
                updates_by_col[col_name] = []

            updates_by_col[col_name].append((u['value'], u['row_id']))

        for col_name, batch in updates_by_col.items():
            self.cursor.executemany(f'UPDATE {self.formatted_table_name} SET "{col_name}" = ? WHERE rowid = ?', batch)

        self.conn.commit()
        return True

    def overwrite_formatted_data(self, rows):
        print(f"DEBUG: overwrite_formatted_data called with {len(rows) if rows else 0} rows")
        
        # Always use the correct 11 headers - don't recover from DB which might have old schema
        self.formatted_headers = ["Sl. No", "Inv No", "Date", "Patient ID", "Patient Name", "Invoice Balance", "Amt To Adjust", "CustomerCode", "Remark 1", "Remark 2", "Remark 3"]

        print(f"DEBUG: overwrite_formatted_data using {len(self.formatted_headers)} headers")

        self._create_formatted_table(self.formatted_headers)
        if rows:
            placeholder = ','.join(['?']*len(self.formatted_headers))
            cols = ', '.join([f'"{h}"' for h in self.formatted_headers])
            print(f"DEBUG: insert with {len(self.formatted_headers)} cols, row has {len(rows[0]) if rows else 0} values")
            self.cursor.executemany(f"INSERT INTO {self.formatted_table_name} ({cols}) VALUES ({placeholder})", rows)
        self.conn.commit()

    def _load_csv(self, path):
        df = pd.read_csv(path)
        self._process_dataframe(df)

    def _load_excel(self, path, sheet_name=None):
        with open(path, 'rb') as f:
            header_check = f.read(512).decode('utf-8', errors='ignore')

        if '<?xml' in header_check[:200] and 'spreadsheet' in header_check.lower():
            df = self._parse_spreadsheetml(path)
        else:
            try:
                if sheet_name:
                    df = pd.read_excel(path, sheet_name=sheet_name)
                else:
                    df = pd.read_excel(path)
            except Exception as e:
                try:
                    dfs = pd.read_html(path, flavor='html.parser')
                    df = dfs[0] if dfs else pd.DataFrame()
                except:
                    raise e

        self._process_dataframe(df)

    def _parse_spreadsheetml(self, path):
        import xml.etree.ElementTree as ET
        try:
            tree = ET.parse(path)
            root = tree.getroot()
        except Exception as e:
            print(f"ET.parse failed: {e}")
            return pd.DataFrame()

        ns = {
            'ss': 'urn:schemas-microsoft-com:office:spreadsheet',
            'x': 'urn:schemas-microsoft-com:office:spreadsheet'
        }

        worksheet = root.find('.//ss:Worksheet', ns)
        if worksheet is None:
            worksheet = root.find('.//x:Worksheet', ns)

        if worksheet is None:
            return pd.DataFrame()

        data = []
        table = worksheet.find('.//ss:Table', ns) or worksheet.find('.//x:Table', ns)
        if table is not None:
            for row_elem in table.findall('.//ss:Row', ns) or table.findall('.//x:Row', ns):
                row_data = {}
                col_idx = 0
                for cell_elem in row_elem.findall('.//ss:Cell', ns) or row_elem.findall('.//x:Cell', ns):
                    index = cell_elem.get('{urn:schemas-microsoft-com:office:spreadsheet}Index')
                    if index:
                        col_idx = int(index) - 1

                    data_elem = cell_elem.find('.//ss:Data', ns) or cell_elem.find('.//x:Data', ns)
                    if data_elem is not None:
                        row_data[col_idx] = data_elem.text
                    col_idx += 1
                data.append(row_data)

        df = pd.DataFrame(data)
        if not df.empty:
            cols = sorted(df.columns)
            df = df.reindex(columns=cols)

        return df

    def _process_dataframe(self, df):
        df = self._detect_and_fix_headers(df)

        for col in df.columns:
            if pd.api.types.is_datetime64_any_dtype(df[col]):
                df[col] = df[col].astype(str).replace('NaT', None)

        self.headers = [str(c).strip() for c in df.columns]
        self.headers = [h if h and h != 'nan' else f"col_{i}" for i, h in enumerate(self.headers)]
        self.headers = self._deduplicate_headers(self.headers)

        self._create_table(self.headers)

        rows = df.astype(object).where(pd.notnull(df), "").values.tolist()
        rows = self._clean_numeric_columns(rows)

        if rows:
            # Pad rows to match the new headers (which include the extra columns)
            num_cols = len(self.headers)
            padded_rows = [list(r) + [""] * (num_cols - len(r)) for r in rows]
            
            placeholder = ','.join(['?']*num_cols)
            cols = ', '.join([f'"{h}"' for h in self.headers])
            self.cursor.executemany(f"INSERT INTO {self.table_name} ({cols}) VALUES ({placeholder})", padded_rows)
        self.conn.commit()

    def _clean_numeric_columns(self, rows):
        cleaned_rows = []
        for row in rows:
            cleaned_row = []
            for val in row:
                if val is None:
                    cleaned_row.append(None)
                elif isinstance(val, str):
                    if re.match(r'^\d{1,3}(,\d{3})*(\.\d+)?$', val.strip()):
                        cleaned_row.append(val.replace(',', ''))
                    else:
                        cleaned_row.append(val)
                else:
                    cleaned_row.append(val)
            cleaned_rows.append(cleaned_row)
        return cleaned_rows

    def _detect_and_fix_headers(self, df):
        keywords = ['invoice', 'date', 'patient', 'name', 'amount', 'balance', 'code', 'member', 'policy', 'rec', 'no']
        best_idx = -1
        max_matches = 0

        current_cols = [str(c).lower() for c in df.columns]
        current_matches = sum(1 for k in keywords if any(k in c for c in current_cols))
        max_matches = current_matches

        for i in range(min(10, len(df))):
            row = df.iloc[i]
            row_str = [str(cell).lower() for cell in row if pd.notnull(cell)]
            matches = sum(1 for k in keywords if any(k in r for r in row_str))

            if matches > max_matches:
                max_matches = matches
                best_idx = i

        if best_idx == -1 and len(df) > 1:
            row1_str = [str(df.iloc[1, c]).lower() for c in range(min(5, len(df.columns)))]
            if any('invoice' in c or 'batch' in c or 'claim' in c for c in row1_str):
                best_idx = 1

        if best_idx != -1 and max_matches > 0 and not any('Col' in str(c) for c in df.columns):
            df.columns = df.iloc[best_idx]
            df = df.iloc[best_idx + 1:]
            df.reset_index(drop=True, inplace=True)

        return df

    def _load_xls(self, path, sheet_name=None):
        self._load_excel(path, sheet_name)

    def _load_pdf(self, path):
        self.headers = ["Page", "Content"]
        self._create_table(self.headers)

        try:
            if os.path.getsize(path) > 5 * 1024 * 1024:  # 5MB limit
                raise ValueError("PDF file too large (max 5MB)")
            
            reader = PdfReader(path)
            
            # Process max 50 pages to avoid memory issues
            max_pages = min(len(reader.pages), 50)
            
            for i in range(max_pages):
                try:
                    page = reader.pages[i]
                    text = page.extract_text()
                    if text:
                        # Take only first 1000 chars per page to save memory
                        text = text[:1000]
                        self.cursor.execute(
                            f"INSERT INTO {self.table_name} (\"Page\", \"Content\") VALUES (?, ?)",
                            (str(i+1), text)
                        )
                except Exception:
                    continue
            
            self.conn.commit()
        except Exception as e:
            print(f"PDF load error: {e}")
            self.headers = []

    def _deduplicate_headers(self, headers):
        seen = {}
        new_headers = []
        for h in headers:
            h = str(h).strip()
            if h in seen:
                seen[h] += 1
                new_headers.append(f"{h}_{seen[h]}")
            else:
                seen[h] = 0
                new_headers.append(h)
        return new_headers

    def _create_table(self, headers):
        # Ensure the required mapping and remark columns are present
        required_cols = ["Invoice Mapping", "Remark 1", "Remark 2", "Remark 3"]
        full_headers = list(headers)
        for rc in required_cols:
            if rc not in full_headers:
                full_headers.append(rc)
        
        cols = [f'"{h}" TEXT' for h in full_headers]
        try:
            self.cursor.execute(f"DROP TABLE IF EXISTS {self.table_name}")
            self.conn.commit()
        except Exception:
            pass

        self.cursor.execute(f"CREATE TABLE IF NOT EXISTS {self.table_name} ({', '.join(cols)})")
        self.conn.commit()
        
        # Update internal headers to include the new columns for UI display
        self.headers = full_headers

    def combine_remarks(self, invoice_col, remark1_col=None, remark2_col=None, remark3_col=None):
        try:
            # 1. Fetch all data
            self.cursor.execute(f'SELECT rowid, * FROM {self.table_name}')
            rows = [dict(r) for r in self.cursor.fetchall()]
            if not rows: return True

            # 2. Group remarks by invoice for each remark column
            from collections import defaultdict
            invoice_remarks1 = defaultdict(list)
            invoice_remarks2 = defaultdict(list)
            invoice_remarks3 = defaultdict(list)
            
            # Group remarks for each column
            for row in rows:
                inv = str(row.get(invoice_col, "")).strip()
                if inv and inv != "nan":
                    # Remark 1
                    if remark1_col:
                        rem1 = str(row.get(remark1_col, "")).strip()
                        if rem1 and rem1 != "nan" and rem1 not in invoice_remarks1[inv]:
                            invoice_remarks1[inv].append(rem1)
                    # Remark 2
                    if remark2_col:
                        rem2 = str(row.get(remark2_col, "")).strip()
                        if rem2 and rem2 != "nan" and rem2 not in invoice_remarks2[inv]:
                            invoice_remarks2[inv].append(rem2)
                    # Remark 3
                    if remark3_col:
                        rem3 = str(row.get(remark3_col, "")).strip()
                        if rem3 and rem3 != "nan" and rem3 not in invoice_remarks3[inv]:
                            invoice_remarks3[inv].append(rem3)

            # 3. Update the first occurrence of each invoice with the combined remarks
            processed_invoices = set()
            for row in rows:
                inv = str(row.get(invoice_col, "")).strip()
                row_id = row['rowid']
                
                if inv and inv not in processed_invoices:
                    # Combine remarks for each column
                    combined_rem1 = " | ".join(invoice_remarks1[inv]) if inv in invoice_remarks1 else ""
                    combined_rem2 = " | ".join(invoice_remarks2[inv]) if inv in invoice_remarks2 else ""
                    combined_rem3 = " | ".join(invoice_remarks3[inv]) if inv in invoice_remarks3 else ""
                    
                    # Update both the audit columns AND the original columns
                    self.cursor.execute(
                        f'UPDATE {self.table_name} SET "Invoice Mapping" = ?, "Remark 1" = ?, "Remark 2" = ?, "Remark 3" = ?, "{invoice_col}" = ? WHERE rowid = ?',
                        (inv, combined_rem1, combined_rem2, combined_rem3, inv, row_id)
                    )
                    processed_invoices.add(inv)
                elif inv:
                    # For duplicate rows, we still map the invoice for consistency
                    self.cursor.execute(
                        f'UPDATE {self.table_name} SET "Invoice Mapping" = ?, "{invoice_col}" = ? WHERE rowid = ?',
                        (inv, inv, row_id)
                    )

            self.conn.commit()
            return True
        except Exception as e:
            print(f"Combine Remarks Error: {e}")
            return False

    def get_preview(self, limit=None):
        try:
            if not self.headers:
                self._recover_raw_headers()

            self.cursor.execute(f"SELECT rowid, * FROM {self.table_name}")
            rows = self.cursor.fetchall()

            rows = [list(r) for r in rows]

            headings = ['_id'] + self.headers

            return {"headers": headings, "rows": rows}
        except Exception:
            return {"headers": [], "rows": []}

    def delete_raw_rows(self, row_ids):
        if not row_ids: return False

        placeholders = ','.join(['?']*len(row_ids))
        self.cursor.execute(f"DELETE FROM {self.table_name} WHERE rowid IN ({placeholders})", row_ids)
        self.conn.commit()
        return True

    def set_raw_header(self, row_id):
        if not self.headers:
             self._recover_raw_headers()

        if self.headers:
            cols_str = ', '.join([f'"{h}"' for h in self.headers])
            self.cursor.execute(f"SELECT {cols_str} FROM {self.table_name} WHERE rowid = ?", (row_id,))
            new_header_row = self.cursor.fetchone()
        else:
            self.cursor.execute(f"SELECT * FROM {self.table_name} WHERE rowid = ?", (row_id,))
            res = self.cursor.fetchone()
            if res: new_header_row = res[1:]
            else: new_header_row = None

        if not new_header_row:
             raise ValueError("Row not found.")

        self.cursor.execute(f"SELECT rowid, * FROM {self.table_name} WHERE rowid > ?", (row_id,))
        raw_results = self.cursor.fetchall()

        remaining_rows = [r[1:] for r in raw_results]

        new_headers = [str(c).strip() if c else f"col_{i}" for i, c in enumerate(new_header_row)]
        new_headers = self._deduplicate_headers(new_headers)

        self.cursor.execute(f"DROP TABLE IF EXISTS {self.table_name}")
        self._create_table(new_headers)
        self.headers = new_headers

        if remaining_rows:
            placeholder = ','.join(['?']*len(new_headers))
            cols = ', '.join([f'"{h}"' for h in new_headers])
            self.cursor.executemany(f"INSERT INTO {self.table_name} ({cols}) VALUES ({placeholder})", remaining_rows)

        self.conn.commit()
        return True

    def close(self):
        self.conn.close()

    def get_unique_values(self, column_name):
        if not column_name:
            return []
        
        try:
            query = f'SELECT DISTINCT "{column_name}" FROM {self.table_name} WHERE "{column_name}" IS NOT NULL'
            self.cursor.execute(query)
            results = self.cursor.fetchall()
            return [str(row[0]) for row in results if row[0] is not None]
        except Exception as e:
            print(f"Error fetching unique values: {e}")
            return []

    def filter_raw_data(self, column_name, filter_value):
        if not column_name:
            return False
        
        try:
            if filter_value is None:
                query = f'DELETE FROM {self.table_name} WHERE "{column_name}" IS NOT NULL'
            else:
                query = f'DELETE FROM {self.table_name} WHERE "{column_name}" != ?'
            
            self.cursor.execute(query, (filter_value,) if filter_value is not None else ())
            self.conn.commit()
            return True
        except Exception as e:
            print(f"Error filtering raw data: {e}")
            return False
